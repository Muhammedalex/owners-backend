"""
Script to create an advanced Excel template for Property Structure data import
This script creates a comprehensive Excel file with all fields needed for importing ownership data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Colors
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
REQUIRED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_excel_template():
    """Create the Excel template file"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_instructions_sheet(wb)
    create_ownership_sheet(wb)
    create_portfolio_sheet(wb)
    create_portfolio_location_sheet(wb)
    create_building_sheet(wb)
    create_building_floor_sheet(wb)
    create_unit_sheet(wb)
    create_unit_specification_sheet(wb)
    
    # Save file
    filename = f"Property_Structure_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Excel template created successfully: {filename}")
    return filename

def create_instructions_sheet(wb):
    """Create instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)
    
    instructions = [
        ["Property Structure Import Template - Instructions"],
        [""],
        ["IMPORTANT NOTES:"],
        ["1. Fill data starting from row 2 (row 1 contains headers)"],
        ["2. Required fields are marked with yellow background"],
        ["3. Optional fields are marked with gray background"],
        ["4. Example rows are provided in blue background - DELETE before import"],
        ["5. Use exact values for dropdown fields"],
        ["6. Maintain referential integrity (IDs must exist in parent sheets)"],
        [""],
        ["SHEET ORDER (Fill in this order):"],
        ["1. Ownership - First create ownership records"],
        ["2. Portfolio - Link to ownership"],
        ["3. PortfolioLocation - Link to portfolio"],
        ["4. Building - Link to portfolio and ownership"],
        ["5. BuildingFloor - Link to building"],
        ["6. Unit - Link to building, floor, and ownership"],
        ["7. UnitSpecification - Link to unit"],
        [""],
        ["FIELD TYPES:"],
        ["- Text: Enter text values"],
        ["- Number: Enter numeric values"],
        ["- Boolean: Enter 'true' or 'false' (lowercase)"],
        ["- Date: Use format YYYY-MM-DD"],
        ["- Decimal: Use dot (.) as decimal separator"],
        [""],
        ["VALIDATION RULES:"],
        ["- Ownership: name, type, ownership_type, city are required"],
        ["- Portfolio: name, code are required (code must be unique per ownership)"],
        ["- Building: name, code, type, portfolio_id are required"],
        ["- BuildingFloor: building_id, number are required (number unique per building)"],
        ["- Unit: building_id, number, type, area are required"],
        ["- UnitSpecification: unit_id, key are required"],
        [""],
        ["RELATIONSHIPS:"],
        ["- Portfolio.ownership_id → Ownership.id"],
        ["- PortfolioLocation.portfolio_id → Portfolio.id"],
        ["- Building.portfolio_id → Portfolio.id"],
        ["- Building.ownership_id → Ownership.id"],
        ["- BuildingFloor.building_id → Building.id"],
        ["- Unit.building_id → Building.id"],
        ["- Unit.floor_id → BuildingFloor.id"],
        ["- Unit.ownership_id → Ownership.id"],
        ["- UnitSpecification.unit_id → Unit.id"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="366092")
            elif value and value.isupper() and ":" in value:
                cell.font = Font(bold=True, size=11)
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 80

def create_ownership_sheet(wb):
    """Create Ownership sheet"""
    ws = wb.create_sheet("Ownership")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("name", "Ownership name (e.g., 'Al-Rashid Real Estate Company')", True),
        ("legal", "Legal/registered name", False),
        ("type", "Ownership type (company, individual, government, etc.)", True),
        ("ownership_type", "Category (real_estate, investment, etc.)", True),
        ("registration", "Registration number (unique)", False),
        ("tax_id", "Tax identification number", False),
        ("street", "Street address", False),
        ("city", "City name", True),
        ("state", "State/Province", False),
        ("country", "Country (default: Saudi Arabia)", False),
        ("zip_code", "Postal/ZIP code", False),
        ("email", "Contact email", False),
        ("phone", "Contact phone", False),
        ("active", "Active status (true/false, default: true)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        # Add description in row 2
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "شركة الراشد العقارية",  # name
        "شركة الراشد العقارية المساهمة",  # legal
        "company",  # type
        "real_estate",  # ownership_type
        "CR-1234567890",  # registration
        "310123456700003",  # tax_id
        "طريق الملك فهد، 123",  # street
        "الرياض",  # city
        "منطقة الرياض",  # state
        "Saudi Arabia",  # country
        "12345",  # zip_code
        "info@alrashid.com",  # email
        "+966501234567",  # phone
        "true",  # active
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation for type
    type_validation = DataValidation(type="list", formula1='"company,individual,government,organization,other"')
    ws.add_data_validation(type_validation)
    type_validation.add(f"C3:C1000")
    
    # Add data validation for ownership_type
    ownership_type_validation = DataValidation(type="list", formula1='"real_estate,investment,development,management,other"')
    ws.add_data_validation(ownership_type_validation)
    ownership_type_validation.add(f"E3:E1000")
    
    # Add data validation for active
    active_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(active_validation)
    active_validation.add(f"O3:O1000")
    
    # Freeze panes
    ws.freeze_panes = "A3"
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

def create_portfolio_sheet(wb):
    """Create Portfolio sheet"""
    ws = wb.create_sheet("Portfolio")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("ownership_id", "Ownership ID (reference to Ownership.id)", True),
        ("parent_id", "Parent Portfolio ID (for nested portfolios, leave empty if root)", False),
        ("name", "Portfolio name", True),
        ("code", "Portfolio code (unique per ownership)", True),
        ("type", "Portfolio type (general, residential, commercial, mixed, industrial)", False),
        ("description", "Portfolio description", False),
        ("area", "Total area in square meters (decimal)", False),
        ("active", "Active status (true/false, default: true)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "1",  # ownership_id
        "",  # parent_id (empty for root)
        "المشاريع السكنية في الرياض",  # name
        "PORT-001-01",  # code
        "residential",  # type
        "محفظة المشاريع السكنية في منطقة الرياض",  # description
        "25000.50",  # area
        "true",  # active
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation
    type_validation = DataValidation(type="list", formula1='"general,residential,commercial,mixed,industrial"')
    ws.add_data_validation(type_validation)
    type_validation.add(f"F3:F1000")
    
    active_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(active_validation)
    active_validation.add(f"I3:I1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

def create_portfolio_location_sheet(wb):
    """Create PortfolioLocation sheet"""
    ws = wb.create_sheet("PortfolioLocation")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("portfolio_id", "Portfolio ID (reference to Portfolio.id)", True),
        ("street", "Street address", False),
        ("city", "City name", False),
        ("state", "State/Province", False),
        ("country", "Country (default: Saudi Arabia)", False),
        ("zip_code", "Postal/ZIP code", False),
        ("latitude", "Latitude coordinate (decimal, -90 to 90)", False),
        ("longitude", "Longitude coordinate (decimal, -180 to 180)", False),
        ("primary", "Primary location flag (true/false, default: false)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "1",  # portfolio_id
        "طريق الملك فهد",  # street
        "الرياض",  # city
        "منطقة الرياض",  # state
        "Saudi Arabia",  # country
        "12345",  # zip_code
        "24.7136",  # latitude
        "46.6753",  # longitude
        "true",  # primary
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation
    primary_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(primary_validation)
    primary_validation.add(f"J3:J1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

def create_building_sheet(wb):
    """Create Building sheet"""
    ws = wb.create_sheet("Building")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("portfolio_id", "Portfolio ID (reference to Portfolio.id)", True),
        ("ownership_id", "Ownership ID (reference to Ownership.id)", True),
        ("parent_id", "Parent Building ID (for nested buildings, leave empty if root)", False),
        ("name", "Building name", True),
        ("code", "Building code (unique per ownership)", True),
        ("type", "Building type (residential, commercial, mixed, office, retail)", True),
        ("description", "Building description", False),
        ("street", "Street address", False),
        ("city", "City name", False),
        ("state", "State/Province", False),
        ("country", "Country (default: Saudi Arabia)", False),
        ("zip_code", "Postal/ZIP code", False),
        ("latitude", "Latitude coordinate (decimal, -90 to 90)", False),
        ("longitude", "Longitude coordinate (decimal, -180 to 180)", False),
        ("floors", "Number of floors (integer, min: 1)", False),
        ("year", "Construction year (integer, 1800 to current year + 10)", False),
        ("active", "Active status (true/false, default: true)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "1",  # portfolio_id
        "1",  # ownership_id
        "",  # parent_id
        "برج الراشد السكني",  # name
        "BLD-001-01",  # code
        "residential",  # type
        "برج سكني فاخر مكون من 10 طوابق",  # description
        "طريق الملك فهد، 123",  # street
        "الرياض",  # city
        "منطقة الرياض",  # state
        "Saudi Arabia",  # country
        "12345",  # zip_code
        "24.7136",  # latitude
        "46.6753",  # longitude
        "10",  # floors
        "2020",  # year
        "true",  # active
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation
    type_validation = DataValidation(type="list", formula1='"residential,commercial,mixed,office,retail,warehouse,industrial"')
    ws.add_data_validation(type_validation)
    type_validation.add(f"G3:G1000")
    
    active_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(active_validation)
    active_validation.add(f"R3:R1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

def create_building_floor_sheet(wb):
    """Create BuildingFloor sheet"""
    ws = wb.create_sheet("BuildingFloor")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("building_id", "Building ID (reference to Building.id)", True),
        ("number", "Floor number (integer, unique per building, can be negative for basements)", True),
        ("name", "Floor name (e.g., 'Ground Floor', 'الطابق الأرضي')", False),
        ("description", "Floor description", False),
        ("units", "Number of units on this floor (integer, min: 0)", False),
        ("active", "Active status (true/false, default: true)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "1",  # building_id
        "1",  # number
        "الطابق الأرضي",  # name
        "الطابق الأول من البرج",  # description
        "5",  # units
        "true",  # active
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation
    active_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(active_validation)
    active_validation.add(f"G3:G1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

def create_unit_sheet(wb):
    """Create Unit sheet"""
    ws = wb.create_sheet("Unit")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("building_id", "Building ID (reference to Building.id)", True),
        ("floor_id", "Floor ID (reference to BuildingFloor.id, optional)", False),
        ("ownership_id", "Ownership ID (reference to Ownership.id)", True),
        ("number", "Unit number (unique per building)", True),
        ("type", "Unit type (apartment, office, shop, warehouse, studio, penthouse)", True),
        ("name", "Unit name", False),
        ("description", "Unit description", False),
        ("area", "Unit area in square meters (decimal, required)", True),
        ("price_monthly", "Monthly price in SAR (decimal)", False),
        ("price_quarterly", "Quarterly price in SAR (decimal)", False),
        ("price_yearly", "Yearly price in SAR (decimal)", False),
        ("status", "Unit status (available, rented, maintenance, reserved, sold)", False),
        ("active", "Active status (true/false, default: true)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data
    example_row = [
        "",  # id
        "1",  # building_id
        "1",  # floor_id
        "1",  # ownership_id
        "0101",  # number
        "apartment",  # type
        "شقة 0101",  # name
        "شقة فاخرة بغرفتين",  # description
        "120.50",  # area
        "6000.00",  # price_monthly
        "17100.00",  # price_quarterly
        "64800.00",  # price_yearly
        "available",  # status
        "true",  # active
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
    
    # Add data validation
    type_validation = DataValidation(type="list", formula1='"apartment,office,shop,warehouse,studio,penthouse"')
    ws.add_data_validation(type_validation)
    type_validation.add(f"F3:F1000")
    
    status_validation = DataValidation(type="list", formula1='"available,rented,maintenance,reserved,sold"')
    ws.add_data_validation(status_validation)
    status_validation.add(f"M3:M1000")
    
    active_validation = DataValidation(type="list", formula1='"true,false"')
    ws.add_data_validation(active_validation)
    active_validation.add(f"N3:N1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

def create_unit_specification_sheet(wb):
    """Create UnitSpecification sheet"""
    ws = wb.create_sheet("UnitSpecification")
    
    headers = [
        ("id", "Auto-generated ID (leave empty)", False),
        ("unit_id", "Unit ID (reference to Unit.id)", True),
        ("key", "Specification key (e.g., 'bedrooms', 'bathrooms', 'parking')", True),
        ("value", "Specification value", False),
        ("type", "Value type (integer, boolean, string)", False),
    ]
    
    # Write headers
    for col_idx, (header, description, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL if not required else PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        desc_cell = ws.cell(row=2, column=col_idx, value=description)
        desc_cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.border = BORDER
    
    # Example data rows
    example_rows = [
        ["", "1", "bedrooms", "2", "integer"],
        ["", "1", "bathrooms", "2", "integer"],
        ["", "1", "balcony", "true", "boolean"],
        ["", "1", "parking", "1", "integer"],
        ["", "1", "furnished", "false", "boolean"],
    ]
    
    for row_idx, example_row in enumerate(example_rows, start=3):
        for col_idx, value in enumerate(example_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = EXAMPLE_FILL
            cell.border = BORDER
    
    # Add data validation
    type_validation = DataValidation(type="list", formula1='"integer,boolean,string"')
    ws.add_data_validation(type_validation)
    type_validation.add(f"E3:E1000")
    
    # Common keys validation
    key_validation = DataValidation(type="list", formula1='"bedrooms,bathrooms,balcony,parking,furnished,capacity,meeting_rooms,storefront,storage,loading_dock,ceiling_height,security"')
    ws.add_data_validation(key_validation)
    key_validation.add(f"C3:C1000")
    
    ws.freeze_panes = "A3"
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

if __name__ == "__main__":
    try:
        create_excel_template()
    except Exception as e:
        print(f"Error creating Excel template: {e}")
        import traceback
        traceback.print_exc()

